import re
import os
import openpyxl
import pandas as pd
import numpy as np
from datetime import datetime, timedelta

class DrainageSimulationModel:
    """
    Modelo para simular lluvia horaria y evaluar excedentes de drenaje
    Compatible con datos históricos reales de lluvia
    """
    
    def __init__(self, excel_file):
        self.excel_file = excel_file
        self.zones_data = {}
        self.load_data()
    
    def load_data(self):
        """Carga datos de todas las hojas del archivo Excel"""
        if not os.path.isfile(self.excel_file):
            raise FileNotFoundError(f"Archivo no encontrado: {self.excel_file}")

        wb = openpyxl.load_workbook(self.excel_file, data_only=True)

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            # Leer ubicación de A1 (formato: lon="..." lat="...")
            location_text = ws['A1'].value
            lat, lon = self.parse_location(location_text)

            # Buscar encabezados "fecha" y "lluvia" en las primeras filas y columnas
            header_row = None
            fecha_col = None
            lluvia_col = None

            max_search_rows = min(10, ws.max_row)
            max_search_cols = min(10, ws.max_column)

            for r in range(1, max_search_rows + 1):
                found_fecha = None
                found_lluvia = None
                for c in range(1, max_search_cols + 1):
                    val = ws.cell(row=r, column=c).value
                    if not val:
                        continue
                    s = str(val).strip().lower()
                    if 'fecha' in s:
                        found_fecha = c
                    if 'lluv' in s:  # detecta 'lluvia', 'lluvia (mm)', etc.
                        found_lluvia = c
                if found_fecha or found_lluvia:
                    header_row = r
                    fecha_col = found_fecha or 1
                    lluvia_col = found_lluvia or 2
                    break

            # Si no se encontraron encabezados, usar valores por defecto (A2/B2)
            if not header_row:
                header_row = 2
                fecha_col = fecha_col or 1
                lluvia_col = lluvia_col or 2

            # Leer datos desde la fila siguiente a los encabezados
            data = []
            for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row, values_only=True):
                # Protección por si las columnas no existen en la fila actual
                fecha_val = row[fecha_col - 1] if fecha_col and len(row) >= fecha_col else None
                lluvia_val = row[lluvia_col - 1] if lluvia_col and len(row) >= lluvia_col else None

                # Ignorar filas vacías
                if fecha_val is None and (lluvia_val is None or str(lluvia_val).strip() == ''):
                    continue

                # Parsear fecha (varios formatos posibles)
                fecha = None
                if isinstance(fecha_val, datetime):
                    fecha = fecha_val
                else:
                    if fecha_val is not None:
                        s = str(fecha_val).strip()
                        parsed = False
                        for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y'):
                            try:
                                fecha = datetime.strptime(s, fmt)
                                parsed = True
                                break
                            except Exception:
                                continue
                        if not parsed:
                            # intentar convertir número de Excel a fecha
                            try:
                                fecha = openpyxl.utils.datetime.from_excel(float(s))
                            except Exception:
                                fecha = None

                # Parsear lluvia a float (si falla -> 0.0)
                try:
                    lluvia = float(lluvia_val) if lluvia_val not in (None, '') else 0.0
                except Exception:
                    lluvia = 0.0

                data.append({'fecha': fecha, 'lluvia_mm': lluvia})

            # Crear DataFrame y estadísticas básicas
            df = pd.DataFrame(data)
            total_days = len(df)
            total_rainfall = float(df['lluvia_mm'].sum()) if total_days > 0 else 0.0
            max_rainfall = float(df['lluvia_mm'].max()) if total_days > 0 else 0.0
            avg_rainfall = float(df['lluvia_mm'].mean()) if total_days > 0 else 0.0

            self.zones_data[sheet_name] = {
                'latitude': lat,
                'longitude': lon,
                'historical_data': df,
                'total_days': total_days,
                'total_rainfall': total_rainfall,
                'max_rainfall': max_rainfall,
                'avg_rainfall': avg_rainfall
            }

        wb.close()
    
    def parse_location(self, location_text):
        """Extrae latitud y longitud del texto"""
        if not location_text:
            return 0.0, 0.0
        
        text = str(location_text)
        
        # Buscar patrón lon="..." lat="..."
        lon_match = re.search(r'lon\s*=\s*["\']([^"\']+)["\']', text, re.IGNORECASE)
        lat_match = re.search(r'lat\s*=\s*["\']([^"\']+)["\']', text, re.IGNORECASE)
        
        if lon_match and lat_match:
            try:
                lon = float(lon_match.group(1))
                lat = float(lat_match.group(1))
                return lat, lon
            except:
                pass
        
        # Buscar patrón "Latitud: X, Longitud: Y"
        lat_lon = re.findall(r'[-+]?\d*\.?\d+', text)
        if len(lat_lon) >= 2:
            return float(lat_lon[0]), float(lat_lon[1])
        
        # Buscar cualquier número
        numbers = re.findall(r'-?\d+\.?\d*', text)
        if len(numbers) >= 2:
            return float(numbers[0]), float(numbers[1])
        
        return 0.0, 0.0
    
    def simulate_rainfall_from_historical(self, zone_name, hours=24, use_historical_pattern=True):
        """
        Simula lluvia horaria basada en patrones históricos o sintéticos
        
        Args:
            zone_name: Nombre de la zona
            hours: Número de horas a simular
            use_historical_pattern: Si True, usa estadísticas de datos históricos
        """
        if zone_name not in self.zones_data:
            raise ValueError(f"Zona '{zone_name}' no encontrada")
        
        zone = self.zones_data[zone_name]
        
        if use_historical_pattern and len(zone['historical_data']) > 0:
            # Usar estadísticas de datos históricos
            historical_rain = zone['historical_data']['lluvia_mm'].values
            historical_rain = historical_rain[historical_rain > 0]  # Solo días con lluvia
            
            if len(historical_rain) > 0:
                mean_rain = historical_rain.mean()
                std_rain = historical_rain.std()
                max_rain = historical_rain.max()
            else:
                mean_rain = 5
                std_rain = 3
                max_rain = 15
        else:
            # Valores por defecto
            mean_rain = 5
            std_rain = 3
            max_rain = 15
        
        # Generar lluvia con distribución gamma
        if std_rain > 0:
            shape = (mean_rain / std_rain) ** 2
            scale = std_rain ** 2 / mean_rain
        else:
            shape = 2
            scale = mean_rain / 2
        
        rainfall = np.random.gamma(shape, scale, hours)
        rainfall = np.clip(rainfall, 0, max_rain * 1.5)
        
        return rainfall
    
    def simulate_rainfall(self, zone_name, hours=24, intensity='moderate'):
        """
        Simula lluvia horaria con intensidad predefinida
        
        Args:
            zone_name: Nombre de la zona
            hours: Número de horas a simular
            intensity: 'light', 'moderate', 'heavy', 'extreme', 'historical'
        """
        if intensity == 'historical':
            return self.simulate_rainfall_from_historical(zone_name, hours, True)
        
        # Patrones de intensidad de lluvia (mm/hora)
        intensity_patterns = {
            'light': {'mean': 2, 'std': 1, 'max': 5},
            'moderate': {'mean': 5, 'std': 3, 'max': 15},
            'heavy': {'mean': 15, 'std': 8, 'max': 40},
            'extreme': {'mean': 30, 'std': 15, 'max': 80}
        }
        
        pattern = intensity_patterns.get(intensity, intensity_patterns['moderate'])
        
        # Generar lluvia con distribución gamma
        shape = (pattern['mean'] / pattern['std']) ** 2
        scale = pattern['std'] ** 2 / pattern['mean']
        
        rainfall = np.random.gamma(shape, scale, hours)
        rainfall = np.clip(rainfall, 0, pattern['max'])
        
        return rainfall
    
    def calculate_drainage_excess(self, zone_name, rainfall_data, 
                                  drainage_capacity=10.0, area_m2=1000):
        """
        Calcula el excedente de agua respecto a la capacidad de drenaje
        """
        results = []
        accumulated_excess = 0
        
        for hour, rainfall in enumerate(rainfall_data):
            # Calcular excedente por hora
            excess = max(0, rainfall - drainage_capacity)
            accumulated_excess += excess
            
            # Volumen de agua en litros
            volume_liters = (rainfall * area_m2) / 1000
            excess_volume = (excess * area_m2) / 1000
            
            results.append({
                'hora': hour + 1,
                'lluvia_mm': round(rainfall, 2),
                'capacidad_drenaje_mm': drainage_capacity,
                'excedente_mm': round(excess, 2),
                'excedente_acumulado_mm': round(accumulated_excess, 2),
                'volumen_agua_litros': round(volume_liters, 2),
                'excedente_volumen_litros': round(excess_volume, 2),
                'estado': self.get_risk_level(excess)
            })
        
        return pd.DataFrame(results)
    
    def get_risk_level(self, excess):
        """Determina nivel de riesgo según excedente"""
        if excess == 0:
            return 'Normal'
        elif excess < 5:
            return 'Precaución'
        elif excess < 15:
            return 'Alerta'
        elif excess < 30:
            return 'Peligro'
        else:
            return 'Emergencia'
    
    def evaluate_scenario(self, zone_name, scenario_config):
        """
        Evalúa un escenario preventivo
        """
        if zone_name not in self.zones_data:
            raise ValueError(f"Zona '{zone_name}' no encontrada")
        
        # Simular lluvia
        rainfall = self.simulate_rainfall(
            zone_name,
            hours=scenario_config.get('hours', 24),
            intensity=scenario_config.get('intensity', 'moderate')
        )
        
        # Calcular excedentes
        results = self.calculate_drainage_excess(
            zone_name,
            rainfall,
            drainage_capacity=scenario_config.get('drainage_capacity', 10),
            area_m2=scenario_config.get('area_m2', 1000)
        )
        
        zone = self.zones_data[zone_name]
        
        # Resumen del escenario
        summary = {
            'zona': zone_name,
            'latitud': zone['latitude'],
            'longitud': zone['longitude'],
            'datos_historicos': {
                'total_dias': zone.get('total_days', 0),
                'lluvia_total_historica': round(zone.get('total_rainfall', 0), 2),
                'lluvia_maxima_historica': round(zone.get('max_rainfall', 0), 2),
                'lluvia_promedio_historica': round(zone.get('avg_rainfall', 0), 2)
            },
            'simulacion': {
                'total_lluvia_mm': round(rainfall.sum(), 2),
                'lluvia_maxima_mm': round(rainfall.max(), 2),
                'excedente_total_mm': round(results['excedente_acumulado_mm'].iloc[-1], 2),
                'horas_con_excedente': len(results[results['excedente_mm'] > 0]),
                'max_nivel_riesgo': results.loc[results['excedente_mm'].idxmax(), 'estado'] if len(results) > 0 else 'Normal',
                'volumen_total_litros': round(results['volumen_agua_litros'].sum(), 2),
                'volumen_excedente_litros': round(results['excedente_volumen_litros'].sum(), 2)
            }
        }
        
        return results, summary
    
    def get_zones_list(self):
        """Retorna lista de zonas disponibles con sus estadísticas"""
        zones_info = []
        for name, data in self.zones_data.items():
            zones_info.append({
                'nombre': name,
                'latitud': data['latitude'],
                'longitud': data['longitude'],
                'dias_de_datos': data.get('total_days', 0),
                'lluvia_total_mm': round(data.get('total_rainfall', 0), 2)
            })
        return zones_info
    
    def export_results(self, results, summary, output_file='resultados_simulacion.xlsx'):
        """Exporta resultados a Excel"""
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            # Hoja de resumen
            summary_flat = {
                'zona': summary['zona'],
                'latitud': summary['latitud'],
                'longitud': summary['longitud'],
                'dias_historicos': summary['datos_historicos']['total_dias'],
                'lluvia_historica_total': summary['datos_historicos']['lluvia_total_historica'],
                'lluvia_historica_max': summary['datos_historicos']['lluvia_maxima_historica'],
                'lluvia_simulada_total': summary['simulacion']['total_lluvia_mm'],
                'excedente_total': summary['simulacion']['excedente_total_mm'],
                'nivel_riesgo_max': summary['simulacion']['max_nivel_riesgo']
            }
            summary_df = pd.DataFrame([summary_flat])
            summary_df.to_excel(writer, sheet_name='Resumen', index=False)
            
            # Hoja de resultados detallados
            results.to_excel(writer, sheet_name='Detalle_Horario', index=False)
            
            # Hoja de datos históricos
            if summary['zona'] in self.zones_data:
                historical = self.zones_data[summary['zona']]['historical_data']
                historical.to_excel(writer, sheet_name='Datos_Historicos', index=False)


# Ejemplo de uso
if __name__ == "__main__":
    # Crear modelo
    modelo = DrainageSimulationModel('datos_zonas.xlsx')
    
    # Listar zonas disponibles
    print("=== ZONAS DISPONIBLES ===")
    for zone in modelo.get_zones_list():
        print(f"\n📍 {zone['nombre']}")
        print(f"   Coordenadas: {zone['latitud']}, {zone['longitud']}")
        print(f"   Días de datos: {zone['dias_de_datos']}")
        print(f"   Lluvia total histórica: {zone['lluvia_total_mm']} mm")
    
    # Configurar escenario
    escenario = {
        'hours': 24,
        'intensity': 'historical',  # Usa patrones históricos
        'drainage_capacity': 8.0,
        'area_m2': 5000
    }
    
    # Evaluar primera zona
    zonas = [z['nombre'] for z in modelo.get_zones_list()]
    if zonas:
        zona = zonas[0]
        resultados, resumen = modelo.evaluate_scenario(zona, escenario)
        
        print(f"\n\n=== RESUMEN DE SIMULACIÓN: {zona} ===")
        print("\n📊 Datos Históricos:")
        for key, value in resumen['datos_historicos'].items():
            print(f"   {key}: {value}")
        
        print("\n🌧️ Simulación:")
        for key, value in resumen['simulacion'].items():
            print(f"   {key}: {value}")
        
        print("\n=== PRIMERAS 6 HORAS ===")
        print(resultados.head(6))
        
        # Exportar resultados
        modelo.export_results(resultados, resumen)