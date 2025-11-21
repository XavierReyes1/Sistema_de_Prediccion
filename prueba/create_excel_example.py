import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import random
from datetime import datetime, timedelta

def create_sample_excel():
    """
    Crea un archivo Excel con el formato exacto:
    - A1: lon="..." lat="..."
    - A2: fecha | B2: Lluvia (mm)
    - A3+: Fechas y datos de lluvia
    """
    wb = openpyxl.Workbook()
    
    # Eliminar hoja por defecto
    wb.remove(wb.active)
    
    # Zonas de ejemplo en Honduras con coordenadas reales
    zones = [
        {
            'name': 'Tegucigalpa Centro',
            'lat': 14.05387498976198,
            'lon': -87.24557944538367
        },
        {
            'name': 'Comayagüela Norte',
            'lat': 14.0844,
            'lon': -87.2067
        },
        {
            'name': 'Colonia Kennedy',
            'lat': 14.1094,
            'lon': -87.1767
        },
        {
            'name': 'El Hatillo',
            'lat': 14.0375,
            'lon': -87.1544
        },
        {
            'name': 'San Pedro Sula',
            'lat': 15.5041,
            'lon': -88.0250
        }
    ]
    
    # Generar fechas desde mayo 2024
    start_date = datetime(2024, 5, 10)
    
    for zone in zones:
        # Crear hoja
        ws = wb.create_sheet(title=zone['name'])
        
        # A1: Ubicación en formato lon="..." lat="..."
        location_text = f'lon="{zone["lon"]}" lat="{zone["lat"]}"'
        ws['A1'] = location_text
        ws['A1'].font = Font(bold=True, size=11)
        
        # A2, B2: Encabezados
        ws['A2'] = "fecha"
        ws['B2'] = "Lluvia (mm)"
        
        # Estilo para encabezados
        for cell in ['A2', 'B2']:
            ws[cell].font = Font(bold=True)
            ws[cell].alignment = Alignment(horizontal="center", vertical="center")
        
        # Generar datos de lluvia (similar a tu imagen: mayoría son 0, algunos valores)
        num_days = random.randint(30, 60)  # Entre 1-2 meses de datos
        
        for i in range(num_days):
            row = 3 + i
            current_date = start_date + timedelta(days=i)
            
            # 80% de días sin lluvia, 20% con lluvia
            if random.random() < 0.2:
                # Días con lluvia: valores entre 0.1 y 30 mm
                rainfall = round(random.choice([
                    random.uniform(0.1, 2),    # Llovizna
                    random.uniform(2, 10),     # Lluvia moderada
                    random.uniform(10, 30)     # Lluvia fuerte
                ]), 1)
            else:
                rainfall = 0
            
            # Escribir fecha y lluvia
            ws[f'A{row}'] = current_date.strftime('%d/%m/%Y')
            ws[f'B{row}'] = rainfall
            
            # Formato de número para lluvia
            ws[f'B{row}'].number_format = '0.0'
        
        # Ajustar anchos de columna
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 12
    
    # Guardar archivo
    filename = 'datos_zonas.xlsx'
    wb.save(filename)
    print(f"✅ Archivo '{filename}' creado exitosamente")
    print(f"📊 Se crearon {len(zones)} hojas con datos históricos de lluvia")
    print("\nEstructura de cada hoja:")
    print('  A1: lon="..." lat="..."')
    print("  A2: fecha | B2: Lluvia (mm)")
    print("  A3+: Fechas y datos diarios de lluvia")
    print(f"\nCada zona tiene entre 30-60 días de datos históricos")
    print("Formato de ejemplo:")
    print("  10/5/2024    0.3")
    print("  11/5/2024    0")
    print("  16/5/2024    22.3")
    
    return filename

if __name__ == "__main__":
    create_sample_excel()